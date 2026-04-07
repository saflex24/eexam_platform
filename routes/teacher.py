from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import current_user, login_required
from extensions import db
from models.exam import Exam, Question, QuestionOption, ExamResult, ExamSession, StudentAnswer
from models.class_model import Class
from models.user import Teacher, User, Student
from utils.decorators import teacher_required
from utils.file_handler import import_questions_from_csv, generate_questions_template, save_upload_file
from datetime import datetime, timedelta
from werkzeug.utils import secure_filename
import pandas as pd
import io
import os
from models.class_model import Class
from utils.decorators import teacher_required
import subprocess
import json
import tempfile
import zipfile
from sqlalchemy.orm import joinedload
from sqlalchemy import func, desc
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

teacher_bp = Blueprint('teacher', __name__)

# Image upload configuration
UPLOAD_FOLDER = 'static/uploads/questions'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def save_question_image(file):
    if file and allowed_file(file.filename):
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        filename = secure_filename(file.filename)
        unique_filename = f"{int(datetime.utcnow().timestamp())}_{filename}"
        filepath = os.path.join(UPLOAD_FOLDER, unique_filename)
        file.save(filepath)
        return unique_filename
    return None

def delete_question_image(filename):
    if filename:
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        if os.path.exists(filepath):
            try:
                os.remove(filepath)
                return True
            except Exception as e:
                print(f"Error deleting image: {e}")
    return False

# ==================== DASHBOARD ====================

@teacher_bp.route('/dashboard')
@login_required
def dashboard():
    if current_user.role.name != 'Teacher':
        flash('Access denied. Teacher privileges required.', 'error')
        return redirect(url_for('main.index'))

    try:
        teacher = Teacher.query.filter_by(user_id=current_user.id).first()
        if not teacher:
            flash('Teacher profile not found.', 'error')
            return redirect(url_for('main.index'))

        exams = Exam.query.filter_by(created_by=current_user.id, is_deleted=False).all()
        total_exams = len(exams)
        published_exams = sum(1 for exam in exams if exam.published)
        unpublished_exams = total_exams - published_exams
        total_questions = sum(exam.total_questions for exam in exams)
        exam_ids = [exam.id for exam in exams]

        if exam_ids:
            unique_student_ids = set()
            results = ExamResult.query.filter(ExamResult.exam_id.in_(exam_ids)).all()
            for result in results:
                unique_student_ids.add(result.student_id)
            total_students = len(unique_student_ids)
            total_attempts = len(results)
            avg_score = round(sum(r.percentage for r in results) / len(results), 2) if results else 0
        else:
            total_students = 0
            total_attempts = 0
            avg_score = 0

        recent_exams = Exam.query.filter_by(
            created_by=current_user.id,
            is_deleted=False
        ).order_by(Exam.created_at.desc()).limit(10).all()

        return render_template('teacher/dashboard.html',
                             total_exams=total_exams,
                             published_exams=published_exams,
                             unpublished_exams=unpublished_exams,
                             total_questions=total_questions,
                             total_students=total_students,
                             total_attempts=total_attempts,
                             avg_score=avg_score,
                             recent_exams=recent_exams)
    except Exception as e:
        print(f"ERROR in teacher dashboard: {str(e)}")
        import traceback
        traceback.print_exc()
        flash(f'Error loading dashboard: {str(e)}', 'error')
        return redirect(url_for('main.index'))

# ==================== EXAM MANAGEMENT ====================

@teacher_bp.route('/exam/<int:exam_id>/debug-questions')
@teacher_required
def debug_questions(exam_id):
    exam = Exam.query.get_or_404(exam_id)
    questions = Question.query.filter_by(exam_id=exam_id).all()
    debug_info = {
        'exam_id': exam_id,
        'exam_title': exam.title,
        'total_questions_in_exam_obj': exam.total_questions,
        'actual_questions_count': len(questions),
        'questions': []
    }
    for q in questions:
        debug_info['questions'].append({
            'id': q.id,
            'text': q.question_text[:50],
            'type': q.question_type,
            'marks': q.marks,
            'order': q.order,
            'options_count': len(q.options)
        })
    return jsonify(debug_info)

@teacher_bp.route('/exams', methods=['GET'])
@teacher_required
def manage_exams():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '', type=str)
    status_filter = request.args.get('status', '', type=str)
    query = Exam.query.filter_by(created_by=current_user.id, is_deleted=False)
    if search:
        query = query.filter(Exam.title.ilike(f'%{search}%'))
    if status_filter == 'published':
        query = query.filter_by(published=True)
    elif status_filter == 'unpublished':
        query = query.filter_by(published=False)
    elif status_filter == 'upcoming':
        query = query.filter(Exam.start_date > datetime.utcnow())
    elif status_filter == 'ongoing':
        query = query.filter(Exam.start_date <= datetime.utcnow(), Exam.end_date >= datetime.utcnow())
    elif status_filter == 'completed':
        query = query.filter(Exam.end_date < datetime.utcnow())
    exams = query.order_by(Exam.created_at.desc()).paginate(page=page, per_page=20)
    return render_template('teacher/exams.html', exams=exams, search=search, status_filter=status_filter)

# ==================== CREATE EXAM ====================

@teacher_bp.route('/exam/create', methods=['GET', 'POST'])
@teacher_required
def create_exam():
    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        description = request.form.get('description', '').strip()
        subject = request.form.get('subject', '').strip()
        class_id = request.form.get('class_id', type=int) or None
        total_marks = request.form.get('total_marks', type=float, default=100)
        pass_marks = request.form.get('pass_marks', type=float, default=40)
        duration_minutes = request.form.get('duration_minutes', type=int, default=60)
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')
        passcode = request.form.get('passcode', '').strip() or None
        if not all([title, subject, start_date, end_date]):
            flash('Title, subject, start date, and end date are required.', 'danger')
            return redirect(url_for('teacher.create_exam'))
        try:
            start_dt = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
            end_dt = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
            if start_dt >= end_dt:
                flash('End date must be after start date.', 'danger')
                return redirect(url_for('teacher.create_exam'))
            exam = Exam(
                title=title,
                code=Exam.generate_code(),
                description=description,
                subject=subject,
                class_id=class_id,
                created_by=current_user.id,
                total_marks=total_marks,
                pass_marks=pass_marks,
                duration_minutes=duration_minutes,
                start_date=start_dt,
                end_date=end_dt,
                passcode=passcode,
                shuffle_questions=request.form.get('shuffle_questions') == 'on',
                shuffle_options=request.form.get('shuffle_options') == 'on',
                show_results_immediately=request.form.get('show_results_immediately') == 'on',
                allow_review=request.form.get('allow_review') == 'on',
                allow_student_view_result=request.form.get('allow_student_view_result') == 'on',
                randomize_per_student=request.form.get('randomize_per_student') == 'on'
            )
            db.session.add(exam)
            db.session.commit()
            if passcode:
                flash(f'Exam created successfully with passcode: {passcode}. Now add questions.', 'success')
            else:
                flash('Exam created successfully. Now add questions.', 'success')
            return redirect(url_for('teacher.edit_exam_questions', exam_id=exam.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating exam: {str(e)}', 'danger')
    classes = Class.query.filter_by(is_active=True).all()
    return render_template('teacher/exam_form.html', exam=None, classes=classes)

# ==================== EDIT EXAM ====================

@teacher_bp.route('/exam/<int:exam_id>/edit', methods=['GET', 'POST'])
@teacher_required
def edit_exam(exam_id):
    exam = Exam.query.get_or_404(exam_id)
    if exam.created_by != current_user.id:
        flash('You do not have permission to edit this exam.', 'danger')
        return redirect(url_for('teacher.manage_exams'))
    if request.method == 'POST':
        exam.title = request.form.get('title', '').strip()
        exam.description = request.form.get('description', '').strip()
        exam.subject = request.form.get('subject', '').strip()
        exam.class_id = request.form.get('class_id', type=int) or None
        exam.total_marks = request.form.get('total_marks', type=float, default=100)
        exam.pass_marks = request.form.get('pass_marks', type=float, default=40)
        exam.duration_minutes = request.form.get('duration_minutes', type=int, default=60)
        exam.allow_student_view_result = request.form.get('allow_student_view_result') == 'on'
        passcode = request.form.get('passcode', '').strip()
        exam.passcode = passcode if passcode else None
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')
        if start_date:
            exam.start_date = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        if end_date:
            exam.end_date = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
        try:
            db.session.commit()
            if passcode:
                flash(f'Exam updated successfully. Passcode: {passcode}', 'success')
            else:
                flash('Exam updated successfully.', 'success')
            return redirect(url_for('teacher.manage_exams'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating exam: {str(e)}', 'danger')
    classes = Class.query.filter_by(is_active=True).all()
    return render_template('teacher/exam_form.html', exam=exam, classes=classes)

# ==================== DELETE, PUBLISH, UNPUBLISH ====================

@teacher_bp.route('/exam/<int:exam_id>/delete', methods=['POST'])
@teacher_required
def delete_exam(exam_id):
    exam = Exam.query.get_or_404(exam_id)
    if exam.created_by != current_user.id:
        flash('You do not have permission to delete this exam.', 'danger')
        return redirect(url_for('teacher.manage_exams'))
    try:
        exam.is_deleted = True
        db.session.commit()
        flash('Exam deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting exam: {str(e)}', 'danger')
    return redirect(url_for('teacher.manage_exams'))

@teacher_bp.route('/exam/<int:exam_id>/publish', methods=['POST'])
@teacher_required
def publish_exam(exam_id):
    exam = Exam.query.get_or_404(exam_id)
    if exam.created_by != current_user.id:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    if exam.total_questions == 0:
        return jsonify({'success': False, 'message': 'Cannot publish exam without questions'}), 400
    try:
        exam.published = True
        exam.published_at = datetime.utcnow()
        db.session.commit()
        return jsonify({'success': True, 'message': 'Exam published successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@teacher_bp.route('/exam/<int:exam_id>/unpublish', methods=['POST'])
@teacher_required
def unpublish_exam(exam_id):
    exam = Exam.query.get_or_404(exam_id)
    if exam.created_by != current_user.id:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    try:
        exam.published = False
        db.session.commit()
        return jsonify({'success': True, 'message': 'Exam unpublished successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

# ==================== STUDENT REPORTS ====================

@teacher_bp.route('/exam/<int:exam_id>/student-reports')
@login_required
@teacher_required
def exam_student_reports(exam_id):
    try:
        exam = Exam.query.get_or_404(exam_id)
        if hasattr(exam, 'created_by'):
            if exam.created_by != current_user.id:
                flash('You do not have permission to view this exam.', 'danger')
                return redirect(url_for('teacher.dashboard'))
        elif hasattr(exam, 'teacher'):
            if exam.teacher.id != current_user.id:
                flash('You do not have permission to view this exam.', 'danger')
                return redirect(url_for('teacher.dashboard'))

        results = ExamResult.query.filter_by(exam_id=exam_id).order_by(ExamResult.percentage.desc()).all()
        students_data = []
        for result in results:
            student = Student.query.filter_by(user_id=result.student_id).first()
            if student:
                students_data.append({
                    'result': result,
                    'student': student,
                    'user': student.user,
                    'class': student.class_info
                })

        if results:
            avg_score = sum(r.percentage for r in results) / len(results)
            passed = sum(1 for r in results if r.is_passed)
            pass_rate = (passed / len(results)) * 100
        else:
            avg_score = 0
            pass_rate = 0

        stats = {
            'total_students': len(results),
            'passed': sum(1 for r in results if r.is_passed),
            'failed': sum(1 for r in results if not r.is_passed),
            'avg_score': round(avg_score, 2),
            'pass_rate': round(pass_rate, 2)
        }

        return render_template('teacher/exam_student_reports.html',
                             exam=exam,
                             students_data=students_data,
                             stats=stats)
    except Exception as e:
        print(f"ERROR in exam_student_reports: {str(e)}")
        import traceback
        traceback.print_exc()
        flash('Error loading student reports.', 'danger')
        return redirect(url_for('teacher.dashboard'))

# ==================== SCHOOL SETTINGS ====================

def get_school_settings():
    """Get school settings for reports"""
    logo_path = 'static/uploads/school_logo.png'
    if not os.path.exists(logo_path):
        logo_path = None
    return {
        'school_name': 'Mawo Schools And Educational Services',
        'school_address': 'Minna, Niger State',
        'school_phone': '+23434609708',
        'school_email': 'Saflex24@gmail.com',
        'school_logo_path': logo_path
    }

# ==================== PDF REPORT GENERATOR ====================

def create_student_exam_report_pdf(exam, student, result, school_settings):
    """Create PDF student exam report - 2-row student info, 2-column questions, max 4 pages"""

    questions = Question.query.filter_by(exam_id=exam.id).order_by(Question.order).all()

    answers = []
    if result.exam_session_id:
        answers = StudentAnswer.query.filter_by(exam_session_id=result.exam_session_id).all()
    if not answers:
        q_ids = [q.id for q in questions]
        answers = StudentAnswer.query.filter(
            StudentAnswer.student_id == result.student_id,
            StudentAnswer.question_id.in_(q_ids)
        ).all()

    answer_dict = {ans.question_id: ans for ans in answers}

    print(f"\n=== REPORT DEBUG ===")
    print(f"Student: {student.user.full_name} | Questions: {len(questions)} | Answers: {len(answers)}")
    print(f"===================\n")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=letter,
        rightMargin=0.45*inch, leftMargin=0.45*inch,
        topMargin=0.35*inch,   bottomMargin=0.35*inch
    )
    elements = []
    styles   = getSampleStyleSheet()

    def S(name, size=9, color='#000000', bold=False,
          align=TA_LEFT, before=0, after=2, indent=0):
        return ParagraphStyle(
            name, parent=styles['Normal'],
            fontSize=size,
            textColor=colors.HexColor(color),
            fontName='Helvetica-Bold' if bold else 'Helvetica',
            alignment=align,
            spaceBefore=before, spaceAfter=after,
            leftIndent=indent
        )

    title_s   = S('T',  16, '#1F4788', True,  TA_CENTER, 0, 4)
    school_s  = S('SC', 13, '#1F4788', True,  TA_CENTER, 0, 2)
    info_s    = S('IN',  7, '#666666', False, TA_CENTER, 0, 2)
    heading_s = S('HD', 10, '#1F4788', True,  TA_LEFT,   3, 3)
    legend_s  = S('LG',  7, '#666666', False, TA_LEFT,   0, 2)
    q_s  = S('QS', 8, '#1F4788', True,  TA_LEFT, 4, 2)
    op_s = S('OS', 7, '#333333', False, TA_LEFT, 0, 0, 4)

    def get_q_image(image_field, idx, max_w=3.3*inch, max_h=1.4*inch):
        if not image_field:
            return None
        image_str = str(image_field).strip()
        if not image_str or image_str.lower() == 'none':
            return None
        candidates = [
            image_str,
            os.path.join('static', 'uploads', 'questions', image_str),
            os.path.join('static', 'uploads', image_str),
            os.path.join(UPLOAD_FOLDER, image_str),
        ]
        img_path = next((c for c in candidates if os.path.exists(c)), None)
        if not img_path:
            print(f"  Q{idx} image not found: '{image_str}'")
            return Paragraph(f"<i>[img: {os.path.basename(image_str)}]</i>",
                             S(f'iph{idx}', 7, '#999999'))
        try:
            img = Image(img_path)
            iw, ih = img.imageWidth, img.imageHeight
            if iw and ih:
                ratio = min(max_w / iw, max_h / ih, 1.0)
                img.drawWidth  = iw * ratio
                img.drawHeight = ih * ratio
            else:
                img.drawWidth, img.drawHeight = max_w, max_h
            img.hAlign = 'LEFT'
            print(f"  Q{idx} image OK: {img_path}")
            return img
        except Exception as e:
            print(f"  Q{idx} image error: {e}")
            return Paragraph("<i>[image error]</i>", S(f'ierr{idx}', 7, '#CC0000'))

    def build_question_cell(idx, question, answer, col_w):
        cell_items = []

        cell_items.append(Paragraph(
            f"<b>Q{idx}.</b> {question.question_text} "
            f"<font color='#888888' size='7'>[{question.marks}mk]</font>",
            q_s
        ))

        img = get_q_image(question.image, idx, max_w=col_w - 0.1*inch, max_h=1.2*inch)
        if img:
            cell_items.append(Spacer(1, 0.02*inch))
            cell_items.append(img)
            cell_items.append(Spacer(1, 0.02*inch))

        options = QuestionOption.query.filter_by(
            question_id=question.id
        ).order_by(QuestionOption.option_label).all()

        q_type = (question.question_type or '').strip().lower()
        is_mcq = bool(options) or q_type in ('multiple_choice', 'mcq', 'objective', 'true_false')

        if is_mcq and options:
            sel_id = int(answer.selected_option_id) if (answer and answer.selected_option_id is not None) else None
            for opt in options:
                is_sel = (sel_id is not None and sel_id == int(opt.id))
                if opt.is_correct and is_sel:
                    bg, fc, icon, note = colors.HexColor('#C8E6C9'), '#1B5E20', '✓', ' (your ans - correct)'
                elif opt.is_correct and not is_sel:
                    bg, fc, icon, note = colors.HexColor('#E8F5E9'), '#2E7D32', '✓', ' (correct)'
                elif is_sel and not opt.is_correct:
                    bg, fc, icon, note = colors.HexColor('#FFCDD2'), '#B71C1C', '✗', ' (your ans - wrong)'
                else:
                    bg, fc, icon, note = colors.white, '#333333', ' ', ''

                opt_tbl = Table([[Paragraph(
                    f"<font color='{fc}'><b>[{icon}]{opt.option_label}.</b> {opt.option_text}<i>{note}</i></font>",
                    S(f'o{idx}_{opt.id}', 7, fc, False, TA_LEFT, 0, 0, 0)
                )]], colWidths=[col_w - 0.08*inch])
                opt_tbl.setStyle(TableStyle([
                    ('BACKGROUND',    (0,0),(0,0), bg),
                    ('LEFTPADDING',   (0,0),(0,0), 4),
                    ('RIGHTPADDING',  (0,0),(0,0), 2),
                    ('TOPPADDING',    (0,0),(0,0), 2),
                    ('BOTTOMPADDING', (0,0),(0,0), 2),
                    ('LINEBELOW',     (0,0),(0,0), 0.2, colors.HexColor('#EEEEEE')),
                ]))
                cell_items.append(opt_tbl)

            if sel_id is None:
                cell_items.append(Paragraph(
                    "<i>— no option selected —</i>",
                    S(f'na{idx}', 7, '#999999', False, TA_LEFT, 0, 1, 4)
                ))
        else:
            theory_text = None
            if answer:
                raw = answer.theory_answer
                if raw and str(raw).strip().lower() != 'none':
                    theory_text = str(raw).strip()
            ans_tbl = Table([[Paragraph(
                theory_text if theory_text else "<i>No answer provided</i>",
                S(f'at{idx}', 7, '#1B5E20' if theory_text else '#999999', False, TA_LEFT, 0, 0, 0)
            )]], colWidths=[col_w - 0.08*inch])
            ans_tbl.setStyle(TableStyle([
                ('BACKGROUND',    (0,0),(0,0),
                 colors.HexColor('#E8F5E9') if theory_text else colors.HexColor('#F5F5F5')),
                ('LEFTPADDING',   (0,0),(0,0), 4),
                ('RIGHTPADDING',  (0,0),(0,0), 2),
                ('TOPPADDING',    (0,0),(0,0), 3),
                ('BOTTOMPADDING', (0,0),(0,0), 3),
                ('BOX',           (0,0),(0,0), 0.4, colors.HexColor('#CCCCCC')),
            ]))
            cell_items.append(ans_tbl)

        is_correct     = answer.is_correct     if answer else False
        marks_obtained = answer.marks_obtained if answer else 0
        rc = '#2E7D32' if is_correct else ('#C62828' if answer else '#888888')
        label = '✓ CORRECT' if is_correct else ('✗ INCORRECT' if answer else '— UNATTEMPTED')
        cell_items.append(Paragraph(
            f"<font color='{rc}'><b>{label}</b></font>"
            f"<font color='#555555'>  {marks_obtained}/{question.marks}mk</font>",
            S(f'r{idx}', 7, '#000000', False, TA_LEFT, 2, 3)
        ))

        return cell_items

    logo_path = school_settings.get('school_logo_path')
    if logo_path and os.path.exists(logo_path):
        try:
            logo = Image(logo_path, width=0.75*inch, height=0.75*inch)
            logo.hAlign = 'CENTER'
            elements.append(logo)
            elements.append(Spacer(1, 0.05*inch))
        except Exception as e:
            print(f"Logo error: {e}")

    elements.append(Paragraph(school_settings['school_name'].upper(), school_s))
    elements.append(Paragraph(school_settings.get('school_address', ''), info_s))
    elements.append(Paragraph(
        f"{school_settings.get('school_phone','')} | {school_settings.get('school_email','')}",
        info_s
    ))
    elements.append(Paragraph("STUDENT EXAMINATION REPORT", title_s))
    elements.append(Spacer(1, 0.08*inch))

    duration = getattr(exam, 'duration_minutes', 'N/A')
    dur_str  = f"{duration} min" if duration != 'N/A' else 'N/A'

    cw = [0.85*inch, 1.55*inch, 0.85*inch, 1.55*inch,
          0.75*inch, 1.0*inch,  0.75*inch, 0.75*inch]

    si = Table([
        ['Student', student.user.full_name,
         'Admission', student.admission_number,
         'Exam', exam.title, '', ''],
        ['Class', student.class_info.name if student.class_info else 'N/A',
         'Subject', exam.subject,
         'Date', result.submitted_at.strftime('%b %d, %Y'),
         'Duration', dur_str],
    ], colWidths=cw)

    si.setStyle(TableStyle([
        ('BACKGROUND',    (0,0),(0,-1), colors.HexColor('#E8F4F8')),
        ('BACKGROUND',    (2,0),(2,-1), colors.HexColor('#E8F4F8')),
        ('BACKGROUND',    (4,0),(4,-1), colors.HexColor('#E8F4F8')),
        ('BACKGROUND',    (6,0),(6,-1), colors.HexColor('#E8F4F8')),
        ('FONTNAME',      (0,0),(0,-1), 'Helvetica-Bold'),
        ('FONTNAME',      (2,0),(2,-1), 'Helvetica-Bold'),
        ('FONTNAME',      (4,0),(4,-1), 'Helvetica-Bold'),
        ('FONTNAME',      (6,0),(6,-1), 'Helvetica-Bold'),
        ('FONTSIZE',      (0,0),(-1,-1), 8),
        ('GRID',          (0,0),(-1,-1), 0.4, colors.HexColor('#CCCCCC')),
        ('VALIGN',        (0,0),(-1,-1), 'MIDDLE'),
        ('SPAN',          (5,0),(7,0)),
        ('TOPPADDING',    (0,0),(-1,-1), 3),
        ('BOTTOMPADDING', (0,0),(-1,-1), 3),
    ]))
    elements.append(si)
    elements.append(Spacer(1, 0.1*inch))

    correct_count   = sum(1 for a in answers if a.is_correct)
    incorrect_count = len(answers) - correct_count
    unattempted     = len(questions) - len(answers)

    perf_tbl = Table([
        ['Marks',      f"{result.marks_obtained}/{result.total_marks}"],
        ['Percentage', f"{result.percentage:.1f}%"],
    ], colWidths=[1.0*inch, 1.1*inch])
    perf_tbl.setStyle(TableStyle([
        ('BACKGROUND',    (0,0),(0,-1), colors.HexColor('#E8F4F8')),
        ('FONTNAME',      (0,0),(-1,-1), 'Helvetica-Bold'),
        ('FONTSIZE',      (0,0),(-1,-1), 8),
        ('GRID',          (0,0),(-1,-1), 0.4, colors.HexColor('#CCCCCC')),
        ('ALIGN',         (1,0),(1,-1), 'CENTER'),
        ('VALIGN',        (0,0),(-1,-1), 'MIDDLE'),
        ('TOPPADDING',    (0,0),(-1,-1), 4),
        ('BOTTOMPADDING', (0,0),(-1,-1), 4),
    ]))

    stats_tbl = Table([
        ['Correct', str(correct_count), 'Incorrect', str(incorrect_count)],
        ['Unattempted', str(unattempted), 'Total', str(len(questions))],
    ], colWidths=[1.1*inch, 0.65*inch, 1.1*inch, 0.65*inch])
    stats_tbl.setStyle(TableStyle([
        ('BACKGROUND',    (0,0),(0,-1), colors.HexColor('#E8F5E9')),
        ('BACKGROUND',    (2,0),(2,-1), colors.HexColor('#FFEBEE')),
        ('FONTNAME',      (0,0),(0,-1), 'Helvetica-Bold'),
        ('FONTNAME',      (2,0),(2,-1), 'Helvetica-Bold'),
        ('FONTSIZE',      (0,0),(-1,-1), 8),
        ('GRID',          (0,0),(-1,-1), 0.4, colors.HexColor('#CCCCCC')),
        ('ALIGN',         (0,0),(-1,-1), 'CENTER'),
        ('VALIGN',        (0,0),(-1,-1), 'MIDDLE'),
        ('TOPPADDING',    (0,0),(-1,-1), 4),
        ('BOTTOMPADDING', (0,0),(-1,-1), 4),
    ]))

    lw = Table([[Paragraph("PERFORMANCE",      heading_s)],
                [Spacer(1, 0.03*inch)], [perf_tbl]], colWidths=[2.3*inch])
    rw = Table([[Paragraph("QUESTION SUMMARY", heading_s)],
                [Spacer(1, 0.03*inch)], [stats_tbl]], colWidths=[3.9*inch])
    for w in (lw, rw):
        w.setStyle(TableStyle([('LEFTPADDING',(0,0),(-1,-1),0),
                                ('RIGHTPADDING',(0,0),(-1,-1),0),
                                ('TOPPADDING',(0,0),(-1,-1),0),
                                ('BOTTOMPADDING',(0,0),(-1,-1),1)]))

    spacer_w = 7.1*inch - 2.3*inch - 3.9*inch
    summary_row = Table([[lw, rw]], colWidths=[2.3*inch, 3.9*inch + spacer_w])
    summary_row.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'TOP'),
                                      ('LEFTPADDING',(0,0),(-1,-1),0),
                                      ('RIGHTPADDING',(0,0),(-1,-1),0),
                                      ('TOPPADDING',(0,0),(-1,-1),0),
                                      ('BOTTOMPADDING',(0,0),(-1,-1),0)]))
    elements.append(summary_row)
    elements.append(Spacer(1, 0.08*inch))

    elements.append(Paragraph("DETAILED QUESTION ANALYSIS", heading_s))
    elements.append(Paragraph(
        "<i>Legend: [✓] correct  [✗] wrong  [ ] not selected</i>",
        legend_s
    ))
    elements.append(Spacer(1, 0.04*inch))

    GAP    = 0.35 * inch
    FULL_W = 7.1 * inch
    COL_W  = (FULL_W - GAP) / 2

    q_cells = []
    for idx, question in enumerate(questions, 1):
        answer = answer_dict.get(question.id)
        cell_flowables = build_question_cell(idx, question, answer, COL_W - 0.08*inch)
        rows = [[item] for item in cell_flowables]
        mini = Table(rows, colWidths=[COL_W - 0.08*inch])
        mini.setStyle(TableStyle([
            ('LEFTPADDING',   (0,0),(-1,-1), 0),
            ('RIGHTPADDING',  (0,0),(-1,-1), 0),
            ('TOPPADDING',    (0,0),(-1,-1), 0),
            ('BOTTOMPADDING', (0,0),(-1,-1), 0),
            ('BOX',           (0,0),(-1,-1), 0.4, colors.HexColor('#E0E0E0')),
            ('BACKGROUND',    (0,0),(-1,-1), colors.HexColor('#FAFAFA')),
        ]))
        q_cells.append(mini)

    for i in range(0, len(q_cells), 2):
        left  = q_cells[i]
        right = q_cells[i+1] if i+1 < len(q_cells) else Paragraph('', S(f'empty{i}'))
        row_tbl = Table(
            [[left, right]],
            colWidths=[COL_W, COL_W],
            hAlign='LEFT'
        )
        row_tbl.setStyle(TableStyle([
            ('VALIGN',        (0,0),(-1,-1), 'TOP'),
            ('TOPPADDING',    (0,0),(-1,-1), 0),
            ('BOTTOMPADDING', (0,0),(-1,-1), 4),
            ('LEFTPADDING',   (0,0),(0,-1), 0),
            ('RIGHTPADDING',  (0,0),(0,-1), GAP),
            ('LEFTPADDING',   (1,0),(1,-1), 4),
            ('RIGHTPADDING',  (1,0),(1,-1), 0),
        ]))
        elements.append(row_tbl)

    doc.build(elements)
    buffer.seek(0)
    return buffer

# ==================== REPORT ROUTES ====================

@teacher_bp.route('/exam/<int:exam_id>/generate-report/<int:student_id>')
@login_required
@teacher_required
def generate_student_report(exam_id, student_id):
    try:
        exam = Exam.query.get_or_404(exam_id)
        if hasattr(exam, 'created_by'):
            if exam.created_by != current_user.id:
                flash('You do not have permission to access this exam.', 'danger')
                return redirect(url_for('teacher.dashboard'))

        student = Student.query.filter_by(user_id=student_id).first_or_404()
        result = ExamResult.query.filter_by(
            exam_id=exam_id,
            student_id=student_id
        ).first_or_404()

        school_settings = get_school_settings()
        pdf_buffer = create_student_exam_report_pdf(exam, student, result, school_settings)
        filename = f"Exam_Report_{student.user.full_name.replace(' ', '_')}_{exam.title.replace(' ', '_')}.pdf"

        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        print(f"ERROR generating report: {str(e)}")
        import traceback
        traceback.print_exc()
        flash('Error generating report.', 'danger')
        return redirect(url_for('teacher.exam_student_reports', exam_id=exam_id))


@teacher_bp.route('/exam/<int:exam_id>/generate-all-reports')
@login_required
@teacher_required
def generate_all_reports(exam_id):
    try:
        exam = Exam.query.get_or_404(exam_id)
        if hasattr(exam, 'created_by'):
            if exam.created_by != current_user.id:
                flash('You do not have permission to access this exam.', 'danger')
                return redirect(url_for('teacher.dashboard'))

        results = ExamResult.query.filter_by(exam_id=exam_id).all()
        if not results:
            flash('No students have taken this exam yet.', 'warning')
            return redirect(url_for('teacher.exam_student_reports', exam_id=exam_id))

        school_settings = get_school_settings()
        temp_dir = tempfile.mkdtemp()
        zip_path = os.path.join(temp_dir, 'exam_reports.zip')

        with zipfile.ZipFile(zip_path, 'w') as zipf:
            for result in results:
                student = Student.query.filter_by(user_id=result.student_id).first()
                if student:
                    try:
                        pdf_buffer = create_student_exam_report_pdf(exam, student, result, school_settings)
                        class_prefix = f"{student.class_info.name}_" if student.class_info else ""
                        filename = f"{class_prefix}{student.user.full_name.replace(' ', '_')}_{student.admission_number}.pdf"
                        zipf.writestr(filename, pdf_buffer.getvalue())
                        print(f"Generated report for {student.user.full_name}")
                    except Exception as e:
                        print(f"Error generating report for {student.user.full_name}: {e}")

        return send_file(
            zip_path,
            mimetype='application/zip',
            as_attachment=True,
            download_name=f"Exam_Reports_{exam.subject}_{exam.title.replace(' ', '_')}.zip"
        )
    except Exception as e:
        print(f"ERROR generating all reports: {str(e)}")
        import traceback
        traceback.print_exc()
        flash('Error generating reports.', 'danger')
        return redirect(url_for('teacher.exam_student_reports', exam_id=exam_id))

# ==================== QUESTION MANAGEMENT ====================

@teacher_bp.route('/exam/<int:exam_id>/questions', methods=['GET'])
@teacher_required
def edit_exam_questions(exam_id):
    exam = Exam.query.get_or_404(exam_id)
    if exam.created_by != current_user.id:
        flash('You do not have permission to edit this exam.', 'danger')
        return redirect(url_for('teacher.manage_exams'))
    questions = Question.query.filter_by(exam_id=exam_id).order_by(Question.order).all()
    print(f"DEBUG: Found {len(questions)} questions for exam {exam_id}")
    return render_template('teacher/exam_questions.html', exam=exam, questions=questions)

@teacher_bp.route('/exam/<int:exam_id>/question/create', methods=['GET', 'POST'])
@teacher_required
def create_question(exam_id):
    exam = Exam.query.get_or_404(exam_id)
    if exam.created_by != current_user.id:
        flash('You do not have permission to edit this exam.', 'danger')
        return redirect(url_for('teacher.manage_exams'))

    if request.method == 'POST':
        question_text = request.form.get('question_text', '').strip()
        question_type = request.form.get('question_type', 'mcq')
        marks = request.form.get('marks', type=float, default=1)
        instructions = request.form.get('instructions', '').strip()

        if not question_text:
            flash('Question text is required.', 'danger')
            return redirect(url_for('teacher.create_question', exam_id=exam_id))

        try:
            order = Question.query.filter_by(exam_id=exam_id).count() + 1
            question = Question(
                exam_id=exam_id,
                question_text=question_text,
                question_type=question_type,
                marks=marks,
                order=order,
                instructions=instructions
            )

            if 'question_image' in request.files:
                file = request.files['question_image']
                if file and file.filename:
                    file.seek(0, os.SEEK_END)
                    file_size = file.tell()
                    file.seek(0)
                    if file_size > MAX_FILE_SIZE:
                        flash('Image file size must be less than 5MB', 'danger')
                        return redirect(url_for('teacher.create_question', exam_id=exam_id))
                    saved_filename = save_question_image(file)
                    if saved_filename:
                        question.image = saved_filename
                    else:
                        flash('Invalid image file format. Allowed: JPG, PNG, GIF', 'warning')

            db.session.add(question)
            db.session.flush()

            if question_type == 'mcq':
                correct_answer = request.form.get('correct_answer', 'A')
                for label in ['A', 'B', 'C', 'D']:
                    option_text = request.form.get(f'option_{label}', '').strip()
                    if option_text:
                        option = QuestionOption(
                            question_id=question.id,
                            option_text=option_text,
                            option_label=label,
                            is_correct=(label == correct_answer)
                        )
                        db.session.add(option)

            elif question_type == 'true_false':
                correct_answer = request.form.get('correct_answer', 'true')
                db.session.add(QuestionOption(question_id=question.id, option_text='True',  option_label='A', is_correct=(correct_answer == 'true')))
                db.session.add(QuestionOption(question_id=question.id, option_text='False', option_label='B', is_correct=(correct_answer == 'false')))

            exam.total_questions = Question.query.filter_by(exam_id=exam_id).count() + 1
            exam.total_marks = sum(q.marks for q in exam.questions) + marks
            db.session.commit()
            flash('Question added successfully.', 'success')
            return redirect(url_for('teacher.edit_exam_questions', exam_id=exam_id))

        except Exception as e:
            db.session.rollback()
            flash(f'Error adding question: {str(e)}', 'danger')
            import traceback
            traceback.print_exc()

    return render_template('teacher/question_form.html', exam=exam, question=None)

@teacher_bp.route('/question/<int:question_id>/edit', methods=['GET', 'POST'])
@teacher_required
def edit_question(question_id):
    question = Question.query.get_or_404(question_id)
    exam = question.exam
    if exam.created_by != current_user.id:
        flash('You do not have permission to edit this question.', 'danger')
        return redirect(url_for('teacher.manage_exams'))

    if request.method == 'POST':
        old_marks = question.marks
        old_image = question.image
        question.question_text = request.form.get('question_text', '').strip()
        question.marks = request.form.get('marks', type=float, default=1)
        question.instructions = request.form.get('instructions', '').strip()

        try:
            if 'question_image' in request.files:
                file = request.files['question_image']
                if file and file.filename:
                    file.seek(0, os.SEEK_END)
                    file_size = file.tell()
                    file.seek(0)
                    if file_size > MAX_FILE_SIZE:
                        flash('Image file size must be less than 5MB', 'danger')
                        return redirect(url_for('teacher.edit_question', question_id=question_id))
                    if old_image:
                        delete_question_image(old_image)
                    saved_filename = save_question_image(file)
                    if saved_filename:
                        question.image = saved_filename
                    else:
                        flash('Invalid image file format. Allowed: JPG, PNG, GIF', 'warning')

            if request.form.get('delete_image') == 'on' and old_image:
                delete_question_image(old_image)
                question.image = None

            if question.question_type == 'mcq':
                correct_answer = request.form.get('correct_answer', 'A')
                for label in ['A', 'B', 'C', 'D']:
                    option_text = request.form.get(f'option_{label}', '').strip()
                    option = QuestionOption.query.filter_by(question_id=question_id, option_label=label).first()
                    if option_text:
                        if not option:
                            option = QuestionOption(question_id=question_id, option_label=label)
                            db.session.add(option)
                        option.option_text = option_text
                        option.is_correct = (label == correct_answer)
                    elif option:
                        db.session.delete(option)

            elif question.question_type == 'true_false':
                correct_answer = request.form.get('correct_answer', 'true')
                for option in question.options:
                    if option.option_text == 'True':
                        option.is_correct = (correct_answer == 'true')
                    else:
                        option.is_correct = (correct_answer == 'false')

            exam.total_marks = exam.total_marks - old_marks + question.marks
            db.session.commit()
            flash('Question updated successfully.', 'success')
            return redirect(url_for('teacher.edit_exam_questions', exam_id=exam.id))

        except Exception as e:
            db.session.rollback()
            flash(f'Error updating question: {str(e)}', 'danger')
            import traceback
            traceback.print_exc()

    return render_template('teacher/question_form.html', exam=exam, question=question)

@teacher_bp.route('/question/<int:question_id>/delete', methods=['POST'])
@teacher_required
def delete_question(question_id):
    question = Question.query.get_or_404(question_id)
    exam = question.exam
    if exam.created_by != current_user.id:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    try:
        marks = question.marks
        image = question.image
        if image:
            delete_question_image(image)
        db.session.delete(question)
        exam.total_questions = Question.query.filter_by(exam_id=exam.id).count() - 1
        exam.total_marks = exam.total_marks - marks
        remaining_questions = Question.query.filter_by(exam_id=exam.id).order_by(Question.order).all()
        for i, q in enumerate(remaining_questions, 1):
            q.order = i
        db.session.commit()
        return jsonify({'success': True, 'message': 'Question deleted successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

# ==================== BULK IMPORT ====================

@teacher_bp.route('/exam/<int:exam_id>/import-questions', methods=['GET', 'POST'])
@teacher_required
def import_questions(exam_id):
    exam = Exam.query.get_or_404(exam_id)
    if exam.created_by != current_user.id:
        flash('You do not have permission to edit this exam.', 'danger')
        return redirect(url_for('teacher.manage_exams'))

    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file selected.', 'danger')
            return redirect(url_for('teacher.import_questions', exam_id=exam_id))

        file = request.files['file']
        if file.filename == '':
            flash('No file selected.', 'danger')
            return redirect(url_for('teacher.import_questions', exam_id=exam_id))

        allowed_extensions = {'.csv', '.xlsx', '.xls'}
        file_ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
        if f'.{file_ext}' not in allowed_extensions:
            flash('Please upload a CSV or Excel file.', 'danger')
            return redirect(url_for('teacher.import_questions', exam_id=exam_id))

        try:
            df = pd.read_csv(file) if file_ext == 'csv' else pd.read_excel(file)
            questions_imported = 0
            errors = []
            required_columns = ['question_text', 'question_type', 'points']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                flash(f'Missing required columns: {", ".join(missing_columns)}', 'danger')
                return redirect(url_for('teacher.import_questions', exam_id=exam_id))

            current_count = Question.query.filter_by(exam_id=exam_id).count()
            type_mapping = {
                'mcq': 'mcq', 'multiple_choice': 'mcq',
                'true_false': 'true_false', 'truefalse': 'true_false',
                'theory': 'theory', 'essay': 'theory',
                'short_answer': 'theory', 'short': 'theory'
            }

            for idx, row in df.iterrows():
                try:
                    question_type = type_mapping.get(str(row['question_type']).strip().lower(), str(row['question_type']).strip().lower())
                    order = current_count + questions_imported + 1
                    question = Question(
                        exam_id=exam_id,
                        question_text=str(row['question_text']).strip(),
                        question_type=question_type,
                        marks=float(row['points']),
                        order=order
                    )
                    db.session.add(question)
                    db.session.flush()

                    if question_type == 'mcq':
                        options_data = []
                        labels = ['A', 'B', 'C', 'D']
                        for i in range(1, 5):
                            option_col = f'option_{i}'
                            if option_col in row and pd.notna(row[option_col]):
                                options_data.append({'label': labels[i-1], 'text': str(row[option_col]).strip()})
                        if len(options_data) < 2:
                            errors.append(f'Row {idx + 2}: MCQ must have at least 2 options')
                            db.session.rollback()
                            continue
                        correct_option = int(row.get('correct_option', 1)) if pd.notna(row.get('correct_option')) else 1
                        for i, opt_data in enumerate(options_data, 1):
                            db.session.add(QuestionOption(question_id=question.id, option_text=opt_data['text'], option_label=opt_data['label'], is_correct=(i == correct_option)))

                    elif question_type == 'true_false':
                        correct_answer = str(row.get('correct_answer', 'true')).strip().lower() if pd.notna(row.get('correct_answer')) else 'true'
                        db.session.add(QuestionOption(question_id=question.id, option_text='True',  option_label='A', is_correct=(correct_answer == 'true')))
                        db.session.add(QuestionOption(question_id=question.id, option_text='False', option_label='B', is_correct=(correct_answer == 'false')))

                    elif question_type != 'theory':
                        errors.append(f'Row {idx + 2}: Invalid question type "{question_type}"')
                        db.session.rollback()
                        continue

                    questions_imported += 1
                except Exception as e:
                    errors.append(f'Row {idx + 2}: {str(e)}')
                    db.session.rollback()
                    continue

            if questions_imported > 0:
                exam.total_questions = Question.query.filter_by(exam_id=exam_id).count()
                exam.total_marks = sum(q.marks for q in exam.questions)
                db.session.commit()
                flash(f'✓ {questions_imported} question(s) imported successfully!', 'success')
                if errors:
                    flash(f'⚠ Warnings: {"; ".join(errors[:5])}', 'warning')
            else:
                flash('No questions were imported.', 'danger')
                if errors:
                    flash(f'Errors: {"; ".join(errors[:5])}', 'danger')

            return redirect(url_for('teacher.edit_exam_questions', exam_id=exam_id))

        except Exception as e:
            db.session.rollback()
            flash(f'Error importing questions: {str(e)}', 'danger')
            import traceback
            traceback.print_exc()
            return redirect(url_for('teacher.import_questions', exam_id=exam_id))

    return render_template('teacher/import_questions.html', exam=exam)

@teacher_bp.route('/download/questions-template')
@teacher_required
def download_questions_template():
    data = {
        'question_text': ['What is $2 + 2$?', 'The Earth revolves around the Sun.', 'Explain the process of photosynthesis in detail.', 'What is the value of $\\frac{a}{b}$ when a=10 and b=2?'],
        'question_type': ['mcq', 'true_false', 'theory', 'mcq'],
        'points': [5, 3, 10, 8],
        'option_1': ['2', '', '', '3'],
        'option_2': ['3', '', '', '4'],
        'option_3': ['4', '', '', '5'],
        'option_4': ['5', '', '', '6'],
        'correct_option': [3, '', '', 3],
        'correct_answer': ['', 'true', '', '']
    }
    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Questions', index=False)
        worksheet = writer.sheets['Questions']
        for column in worksheet.columns:
            max_length = max((len(str(cell.value)) for cell in column if cell.value), default=0)
            worksheet.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='questions_template.xlsx')

# ==================== RESULTS ====================

@teacher_bp.route('/exam/<int:exam_id>/results')
@teacher_required
def exam_results(exam_id):
    try:
        print(f"\n=== EXAM RESULTS PAGE ===")
        exam = Exam.query.get_or_404(exam_id)
        if exam.created_by != current_user.id:
            flash('You do not have permission to view these results.', 'danger')
            return redirect(url_for('teacher.dashboard'))

        page     = request.args.get('page', 1, type=int)
        search   = request.args.get('search', '', type=str)
        sort_by  = request.args.get('sort', 'submitted_at', type=str)
        query    = ExamResult.query.filter_by(exam_id=exam_id)

        if search:
            query = query.join(User, ExamResult.student_id == User.id).filter(
                db.or_(User.first_name.ilike(f'%{search}%'), User.last_name.ilike(f'%{search}%'))
            )

        if sort_by == 'marks_obtained':
            query = query.order_by(ExamResult.marks_obtained.desc())
        elif sort_by == 'percentage':
            query = query.order_by(ExamResult.percentage.desc())
        else:
            query = query.order_by(ExamResult.submitted_at.desc())

        results     = query.paginate(page=page, per_page=20, error_out=False)
        all_results = ExamResult.query.filter_by(exam_id=exam_id).all()
        stats = {
            'total_attempts': len(all_results),
            'avg_score':  sum(r.percentage for r in all_results) / len(all_results) if all_results else 0,
            'pass_count': sum(1 for r in all_results if r.is_passed),
            'fail_count': sum(1 for r in all_results if not r.is_passed),
        }

        return render_template('teacher/exam_results.html',
                             exam=exam, results=results, stats=stats,
                             search=search, sort_by=sort_by)
    except Exception as e:
        print(f"ERROR in exam_results: {str(e)}")
        import traceback
        traceback.print_exc()
        flash(f'Error loading exam results: {str(e)}', 'danger')
        return redirect(url_for('teacher.dashboard'))

@teacher_bp.route('/result/<int:result_id>/details')
@teacher_required
def result_details(result_id):
    try:
        result = ExamResult.query.get_or_404(result_id)
        if result.exam.created_by != current_user.id:
            flash('You do not have permission to view this result.', 'danger')
            return redirect(url_for('teacher.dashboard'))

        student_user    = User.query.get(result.student_id)
        student_profile = Student.query.filter_by(user_id=result.student_id).first()
        answers         = StudentAnswer.query.filter_by(exam_session_id=result.exam_session_id).all()
        all_questions   = Question.query.filter_by(exam_id=result.exam_id).all()
        total_questions = len(all_questions)
        answered_count  = len(answers)
        unattempted_count = total_questions - answered_count
        correct_count   = sum(1 for a in answers if a.is_correct == True)
        incorrect_count = sum(1 for a in answers if a.is_correct == False)
        exam_session    = ExamSession.query.get(result.exam_session_id)

        result.correct_answers   = correct_count
        result.incorrect_answers = incorrect_count
        result.unattempted       = unattempted_count

        return render_template('teacher/result_details.html',
                             result=result,
                             student_user=student_user,
                             student_profile=student_profile,
                             answers=answers,
                             exam_session=exam_session,
                             correct_count=correct_count,
                             incorrect_count=incorrect_count,
                             unattempted_count=unattempted_count)
    except Exception as e:
        print(f"ERROR in result_details: {str(e)}")
        import traceback
        traceback.print_exc()
        flash(f'Error loading result details: {str(e)}', 'danger')
        return redirect(url_for('teacher.dashboard'))

@teacher_bp.route('/exam/<int:exam_id>/toggle-result-visibility', methods=['POST'])
@teacher_required
def toggle_result_visibility(exam_id):
    try:
        exam = Exam.query.get_or_404(exam_id)
        if exam.created_by != current_user.id:
            flash('You do not have permission to modify this exam.', 'danger')
            return redirect(url_for('teacher.dashboard'))
        current_status = getattr(exam, 'allow_student_view_result', True)
        exam.allow_student_view_result = not current_status
        db.session.commit()
        status = "visible to students" if exam.allow_student_view_result else "hidden from students"
        flash(f'Exam results are now {status}!', 'success')
        return redirect(url_for('teacher.exam_results', exam_id=exam_id))
    except Exception as e:
        print(f"Error toggling result visibility: {str(e)}")
        import traceback
        traceback.print_exc()
        flash('Error updating result visibility.', 'danger')
        return redirect(url_for('teacher.dashboard'))

# ==================== ANALYTICS ====================

@teacher_bp.route('/analytics/performance')
@login_required
@teacher_required
def analytics_performance():
    try:
        teacher = Teacher.query.filter_by(user_id=current_user.id).first_or_404()

        # FIX 1: Safely get teacher subject with case-insensitive handling
        teacher_subject = None
        if hasattr(teacher, 'subject') and teacher.subject:
            teacher_subject = teacher.subject.strip()

        # Debug logging to track data flow
        print(f"\n=== ANALYTICS DEBUG ===")
        print(f"Teacher: {current_user.id} | Subject: '{teacher_subject}'")

        exam_id   = request.args.get('exam_id', type=int)
        class_id  = request.args.get('class_id', type=int)
        date_from = request.args.get('date_from')
        date_to   = request.args.get('date_to')

        # FIX 2: Query exams without strict subject filtering if no subject is set
        exams_query = Exam.query.filter_by(created_by=current_user.id, published=True, is_deleted=False)
        if teacher_subject:
            # Use case-insensitive subject matching
            exams_query = exams_query.filter(
                func.lower(Exam.subject) == func.lower(teacher_subject)
            )
        teacher_exams = exams_query.order_by(Exam.created_at.desc()).all()

        print(f"Teacher exams found: {len(teacher_exams)}")

        subjects = [teacher_subject] if teacher_subject else list(
            set(e.subject for e in teacher_exams if e.subject)
        )
        teacher_exam_ids = [e.id for e in teacher_exams]

        if not teacher_exam_ids:
            print("No exam IDs found — returning empty dashboard")
            return render_empty_dashboard(teacher, teacher_subject, subjects)

        # Get all classes that have students who took these exams
        teacher_classes = db.session.query(Class).join(
            Student, Class.id == Student.class_id
        ).join(
            ExamResult, Student.user_id == ExamResult.student_id
        ).filter(
            ExamResult.exam_id.in_(teacher_exam_ids)
        ).distinct().all()

        # Build filtered results query
        results_query = db.session.query(ExamResult).filter(
            ExamResult.exam_id.in_(teacher_exam_ids)
        )
        if exam_id:
            results_query = results_query.filter(ExamResult.exam_id == exam_id)
        if class_id:
            results_query = results_query.join(
                Student, ExamResult.student_id == Student.user_id
            ).filter(Student.class_id == class_id)
        if date_from:
            try:
                results_query = results_query.filter(
                    ExamResult.submitted_at >= datetime.strptime(date_from, '%Y-%m-%d')
                )
            except ValueError:
                pass
        if date_to:
            try:
                results_query = results_query.filter(
                    ExamResult.submitted_at <= datetime.strptime(date_to, '%Y-%m-%d')
                )
            except ValueError:
                pass

        all_results = results_query.all()
        print(f"All results count: {len(all_results)}")

        # Compute all analytics components
        stats               = calculate_teacher_stats(all_results)
        top_performers      = get_top_performers(all_results, 10)
        struggling_students = get_struggling_students(all_results, 10)
        subject_performance = get_subject_performance_for_teacher(teacher_exam_ids, class_id, date_from, date_to)
        exam_performance    = get_exam_performance_for_teacher(teacher_exam_ids, class_id, date_from, date_to)
        class_comparison    = get_class_comparison_for_teacher(teacher_exam_ids, date_from, date_to) if not class_id else []
        grade_distribution  = get_grade_distribution(all_results)
        performance_trend   = get_performance_trend_for_teacher(teacher_exam_ids, class_id)

        # FIX 3: Build enriched recent_submissions list so the template can access
        # student name and exam title without relying on ORM relationships that
        # may not be defined on ExamResult
        raw_recent = results_query.order_by(ExamResult.submitted_at.desc()).limit(15).all()
        recent_submissions = _enrich_recent_submissions(raw_recent)

        charts_data = {
            'grade_dist':   grade_distribution,
            'trend':        performance_trend,
            'subject_perf': subject_performance,
            'class_comp':   class_comparison
        }

        print(f"Stats: {stats}")
        print(f"======================\n")

        return render_template(
            'teacher/analytics_performance.html',
            teacher=teacher,
            stats=stats,
            top_performers=top_performers,
            struggling_students=struggling_students,
            subject_performance=subject_performance,
            exam_performance=exam_performance,
            class_comparison=class_comparison,
            grade_distribution=grade_distribution,
            recent_submissions=recent_submissions,
            performance_trend=performance_trend,
            charts_data=charts_data,
            teacher_exams=teacher_exams,
            subjects=subjects,
            teacher_classes=teacher_classes,
            teacher_subject=teacher_subject,
            filters={
                'exam_id':   exam_id,
                'class_id':  class_id,
                'date_from': date_from,
                'date_to':   date_to
            }
        )
    except Exception as e:
        print(f"Error in analytics_performance: {str(e)}")
        import traceback
        traceback.print_exc()
        flash(f'Error loading analytics: {str(e)}', 'danger')
        return redirect(url_for('teacher.dashboard'))


def _enrich_recent_submissions(raw_results):
    """
    FIX 3 HELPER: Convert raw ExamResult rows into enriched dicts that the
    template can safely access.  Replaces the previous pattern of accessing
    result.student (which is not a mapped relationship on ExamResult).
    """
    enriched = []
    for r in raw_results:
        # Look up the student User row
        student_user = User.query.get(r.student_id)

        # Look up the exam — use the mapped relationship if it exists,
        # otherwise fall back to a direct query so we never crash
        exam_obj = None
        if hasattr(r, 'exam') and r.exam is not None:
            exam_obj = r.exam
        else:
            exam_obj = Exam.query.get(r.exam_id)

        enriched.append({
            'result':       r,
            'student':      student_user,   # User object (has .full_name)
            'exam':         exam_obj,        # Exam object (has .title)
        })
    return enriched

# ==================== ANALYTICS HELPERS ====================

def render_empty_dashboard(teacher, teacher_subject, subjects):
    """Return the analytics template with zero-state data."""
    empty_stats = {
        'total_students':    0,
        'total_submissions': 0,
        'average_score':     0,
        'pass_rate':         0,
        'highest_score':     0,
        'lowest_score':      0,
    }
    empty_grade_dist = {'A': 0, 'B': 0, 'C': 0, 'D': 0, 'F': 0}
    return render_template(
        'teacher/analytics_performance.html',
        teacher=teacher,
        stats=empty_stats,
        top_performers=[],
        struggling_students=[],
        subject_performance=[],
        exam_performance=[],
        class_comparison=[],
        grade_distribution=empty_grade_dist,
        recent_submissions=[],
        performance_trend=[],
        charts_data={
            'grade_dist':   empty_grade_dist,
            'trend':        [],
            'subject_perf': [],
            'class_comp':   []
        },
        teacher_exams=[],
        subjects=subjects,
        teacher_classes=[],
        teacher_subject=teacher_subject,
        filters={
            'exam_id':   None,
            'class_id':  None,
            'date_from': None,
            'date_to':   None
        }
    )


def calculate_teacher_stats(results):
    """Aggregate summary statistics from a list of ExamResult objects."""
    if not results:
        return {
            'total_students':    0,
            'total_submissions': 0,
            'average_score':     0,
            'pass_rate':         0,
            'highest_score':     0,
            'lowest_score':      0,
        }
    total_submissions = len(results)
    unique_students   = len(set(r.student_id for r in results))
    average_score     = sum(r.percentage for r in results) / total_submissions
    pass_rate         = (sum(1 for r in results if r.is_passed) / total_submissions) * 100
    scores            = [r.percentage for r in results]
    return {
        'total_students':    unique_students,
        'total_submissions': total_submissions,
        'average_score':     round(average_score, 2),
        'pass_rate':         round(pass_rate, 2),
        'highest_score':     round(max(scores), 2),
        'lowest_score':      round(min(scores), 2),
    }


def get_top_performers(results, limit=10):
    """
    Return top-performing students sorted by average score descending.

    FIX 4: The returned dicts use 'student' key pointing to a User object
    (not a Student profile) — matching what the template expects:
        perf.student.full_name
        perf.average_score
        perf.total_exams
    """
    if not results:
        return []

    student_scores = {}
    for r in results:
        student_scores.setdefault(r.student_id, []).append(r.percentage)

    performers = []
    for student_id, scores in student_scores.items():
        user    = User.query.get(student_id)
        profile = Student.query.filter_by(user_id=student_id).first()
        if user:
            avg = sum(scores) / len(scores)
            performers.append({
                'student':         user,      # User object — template uses .full_name
                'student_profile': profile,
                'average_score':   round(avg, 2),
                'total_exams':     len(scores),
                'highest_score':   round(max(scores), 2),
                'lowest_score':    round(min(scores), 2),
            })

    performers.sort(key=lambda x: x['average_score'], reverse=True)
    return performers[:limit]


def get_struggling_students(results, limit=10):
    """
    Return students with average score below 50%, sorted ascending.

    FIX 5: Same 'student' → User object convention as get_top_performers so
    both tables in the template use the same access pattern.
    """
    if not results:
        return []

    student_scores = {}
    for r in results:
        student_scores.setdefault(r.student_id, []).append(r.percentage)

    struggling = []
    for student_id, scores in student_scores.items():
        avg = sum(scores) / len(scores)
        if avg < 50:
            user    = User.query.get(student_id)
            profile = Student.query.filter_by(user_id=student_id).first()
            if user:
                struggling.append({
                    'student':         user,    # User object — template uses .full_name
                    'student_profile': profile,
                    'average_score':   round(avg, 2),
                    'total_exams':     len(scores),
                    'failed_exams':    sum(1 for s in scores if s < 50),
                    'lowest_score':    round(min(scores), 2),
                })

    struggling.sort(key=lambda x: x['average_score'])
    return struggling[:limit]


def get_subject_performance_for_teacher(exam_ids, class_id, date_from, date_to):
    if not exam_ids:
        return []
    query = db.session.query(
        Exam.subject,
        func.count(ExamResult.id).label('total'),
        func.avg(ExamResult.percentage).label('avg_score'),
        func.sum(func.cast(ExamResult.is_passed, db.Integer)).label('passed')
    ).join(ExamResult, Exam.id == ExamResult.exam_id).filter(Exam.id.in_(exam_ids))
    if class_id:
        query = query.join(Student, ExamResult.student_id == Student.user_id).filter(
            Student.class_id == class_id
        )
    if date_from:
        try:
            query = query.filter(ExamResult.submitted_at >= datetime.strptime(date_from, '%Y-%m-%d'))
        except ValueError:
            pass
    if date_to:
        try:
            query = query.filter(ExamResult.submitted_at <= datetime.strptime(date_to, '%Y-%m-%d'))
        except ValueError:
            pass
    results = query.group_by(Exam.subject).all()
    return [
        {
            'subject':           subj or 'General',
            'total_submissions': total,
            'average_score':     round(avg, 2) if avg else 0,
            'pass_rate':         round((passed / total * 100), 2) if total > 0 else 0,
        }
        for subj, total, avg, passed in results
    ]


def get_exam_performance_for_teacher(exam_ids, class_id, date_from, date_to):
    if not exam_ids:
        return []
    query = db.session.query(
        Exam,
        func.count(ExamResult.id).label('total'),
        func.avg(ExamResult.percentage).label('avg_score')
    ).join(ExamResult, Exam.id == ExamResult.exam_id).filter(Exam.id.in_(exam_ids))
    if class_id:
        query = query.join(Student, ExamResult.student_id == Student.user_id).filter(
            Student.class_id == class_id
        )
    if date_from:
        try:
            query = query.filter(ExamResult.submitted_at >= datetime.strptime(date_from, '%Y-%m-%d'))
        except ValueError:
            pass
    if date_to:
        try:
            query = query.filter(ExamResult.submitted_at <= datetime.strptime(date_to, '%Y-%m-%d'))
        except ValueError:
            pass
    results = query.group_by(Exam.id).order_by(desc('avg_score')).limit(10).all()
    return [
        {
            'exam':              exam,
            'total_submissions': total,
            'average_score':     round(avg, 2) if avg else 0,
        }
        for exam, total, avg in results
    ]


def get_class_comparison_for_teacher(exam_ids, date_from, date_to):
    if not exam_ids:
        return []
    query = db.session.query(
        Class.name,
        func.count(ExamResult.id).label('total'),
        func.avg(ExamResult.percentage).label('avg_score')
    ).join(Student, Class.id == Student.class_id).join(
        ExamResult, Student.user_id == ExamResult.student_id
    ).filter(ExamResult.exam_id.in_(exam_ids))
    if date_from:
        try:
            query = query.filter(ExamResult.submitted_at >= datetime.strptime(date_from, '%Y-%m-%d'))
        except ValueError:
            pass
    if date_to:
        try:
            query = query.filter(ExamResult.submitted_at <= datetime.strptime(date_to, '%Y-%m-%d'))
        except ValueError:
            pass
    results = query.group_by(Class.id, Class.name).all()
    data = [
        {
            'class_name':        name,
            'total_submissions': total,
            'average_score':     round(avg, 2) if avg else 0,
        }
        for name, total, avg in results
    ]
    data.sort(key=lambda x: x['average_score'], reverse=True)
    return data


def get_grade_distribution(results):
    """
    FIX 6: Fall back to computing grade from percentage when result.grade
    is not set, so the pie chart always has real data.
    """
    dist = {'A': 0, 'B': 0, 'C': 0, 'D': 0, 'F': 0}
    if not results:
        return dist

    for r in results:
        grade = r.grade if (r.grade and r.grade in dist) else _compute_grade(r.percentage)
        dist[grade] += 1

    return dist


def _compute_grade(percentage):
    """Derive a letter grade from a percentage score."""
    if percentage >= 70:
        return 'A'
    elif percentage >= 60:
        return 'B'
    elif percentage >= 50:
        return 'C'
    elif percentage >= 40:
        return 'D'
    else:
        return 'F'


def get_performance_trend_for_teacher(exam_ids, class_id):
    if not exam_ids:
        return []
    six_months_ago = datetime.now() - timedelta(days=180)
    dialect_name   = db.engine.dialect.name
    if dialect_name == 'postgresql':
        month_expr = func.to_char(ExamResult.submitted_at, 'YYYY-MM').label('month')
    elif dialect_name == 'mysql':
        month_expr = func.date_format(ExamResult.submitted_at, '%Y-%m').label('month')
    else:
        # SQLite fallback
        month_expr = func.strftime('%Y-%m', ExamResult.submitted_at).label('month')

    query = db.session.query(
        month_expr,
        func.avg(ExamResult.percentage).label('avg')
    ).filter(
        ExamResult.exam_id.in_(exam_ids),
        ExamResult.submitted_at >= six_months_ago
    )
    if class_id:
        query = query.join(Student, ExamResult.student_id == Student.user_id).filter(
            Student.class_id == class_id
        )
    results = query.group_by(month_expr).order_by(month_expr).all()
    return [
        {'month': m, 'average_score': round(a, 2) if a else 0}
        for m, a in results
    ]

# ==================== EXPORT ====================

@teacher_bp.route('/analytics/export/performance')
@login_required
@teacher_required
def export_performance_report():
    try:
        teacher = Teacher.query.filter_by(user_id=current_user.id).first_or_404()
        teacher_subject = None
        if hasattr(teacher, 'subject') and teacher.subject:
            teacher_subject = teacher.subject.strip()

        exam_id   = request.args.get('exam_id', type=int)
        class_id  = request.args.get('class_id', type=int)
        date_from = request.args.get('date_from')
        date_to   = request.args.get('date_to')

        exams_query = Exam.query.filter_by(created_by=current_user.id, published=True, is_deleted=False)
        if teacher_subject:
            exams_query = exams_query.filter(
                func.lower(Exam.subject) == func.lower(teacher_subject)
            )
        teacher_exams = exams_query.all()
        exam_ids = [e.id for e in teacher_exams]

        if not exam_ids:
            flash('No exam data available to export.', 'warning')
            return redirect(url_for('teacher.analytics_performance'))

        query = db.session.query(ExamResult).filter(ExamResult.exam_id.in_(exam_ids))
        if exam_id:
            query = query.filter(ExamResult.exam_id == exam_id)
        if class_id:
            query = query.join(Student, ExamResult.student_id == Student.user_id).filter(
                Student.class_id == class_id
            )
        if date_from:
            try:
                query = query.filter(ExamResult.submitted_at >= datetime.strptime(date_from, '%Y-%m-%d'))
            except ValueError:
                pass
        if date_to:
            try:
                query = query.filter(ExamResult.submitted_at <= datetime.strptime(date_to, '%Y-%m-%d'))
            except ValueError:
                pass

        results = query.order_by(ExamResult.submitted_at.desc()).all()
        if not results:
            flash('No data to export.', 'warning')
            return redirect(url_for('teacher.analytics_performance'))

        wb = Workbook()
        ws = wb.active
        ws.title = "Performance Report"
        ws.merge_cells('A1:I1')
        title = f"Student Performance Report - {current_user.full_name}"
        if teacher_subject:
            title += f" ({teacher_subject})"
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=14, color="4472C4")
        ws['A1'].alignment = Alignment(horizontal="center")

        headers = ['Student', 'Exam', 'Subject', 'Score%', 'Grade', 'Status', 'Class', 'Date', 'Marks']
        for col, h in enumerate(headers, 1):
            ws.cell(3, col).value = h
            ws.cell(3, col).font = Font(bold=True, color="FFFFFF")
            ws.cell(3, col).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        for row, result in enumerate(results, 4):
            student         = User.query.get(result.student_id)
            student_profile = Student.query.filter_by(user_id=result.student_id).first()
            # FIX 7: Safely get exam object for export rows
            exam_obj = result.exam if hasattr(result, 'exam') and result.exam else Exam.query.get(result.exam_id)

            ws.cell(row, 1).value = student.full_name if student else 'Unknown'
            ws.cell(row, 2).value = exam_obj.title if exam_obj else 'N/A'
            ws.cell(row, 3).value = exam_obj.subject if exam_obj else 'General'
            ws.cell(row, 4).value = round(result.percentage, 2)
            ws.cell(row, 5).value = result.grade or _compute_grade(result.percentage)
            ws.cell(row, 6).value = 'PASS' if result.is_passed else 'FAIL'
            ws.cell(row, 7).value = (
                student_profile.class_info.name
                if student_profile and student_profile.class_info else 'N/A'
            )
            ws.cell(row, 8).value = result.submitted_at.strftime('%Y-%m-%d')
            ws.cell(row, 9).value = f"{result.marks_obtained}/{result.total_marks}"

            if result.is_passed:
                ws.cell(row, 6).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                ws.cell(row, 6).font = Font(color="006100", bold=True)
            else:
                ws.cell(row, 6).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                ws.cell(row, 6).font = Font(color="9C0006", bold=True)

        temp_dir = tempfile.gettempdir()
        filename = f"performance_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(temp_dir, filename)
        wb.save(filepath)
        return send_file(filepath, as_attachment=True, download_name=filename)

    except Exception as e:
        print(f"Export error: {e}")
        import traceback
        traceback.print_exc()
        flash(f'Export error: {str(e)}', 'danger')
        return redirect(url_for('teacher.analytics_performance'))

# ==================== PERFORMANCE CHARTS ====================

@teacher_bp.route('/exam/<int:exam_id>/performance-chart')
@teacher_required
def performance_chart(exam_id):
    exam = Exam.query.get_or_404(exam_id)
    if exam.created_by != current_user.id:
        flash('You do not have permission to view this.', 'danger')
        return redirect(url_for('teacher.manage_exams'))
    results = ExamResult.query.filter_by(exam_id=exam_id).all()
    distributions = {
        '0-20':   len([r for r in results if 0  <= r.percentage < 20]),
        '20-40':  len([r for r in results if 20 <= r.percentage < 40]),
        '40-60':  len([r for r in results if 40 <= r.percentage < 60]),
        '60-80':  len([r for r in results if 60 <= r.percentage < 80]),
        '80-100': len([r for r in results if 80 <= r.percentage <= 100])
    }
    return render_template('teacher/performance_chart.html', exam=exam, results=results, distributions=distributions)

@teacher_bp.route('/api/exam/<int:exam_id>/performance-data')
@teacher_required
def performance_data(exam_id):
    exam = Exam.query.get_or_404(exam_id)
    if exam.created_by != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
    results = ExamResult.query.filter_by(exam_id=exam_id).all()
    return jsonify({
        'labels': ['0-20', '20-40', '40-60', '60-80', '80-100'],
        'data': [
            len([r for r in results if 0  <= r.percentage < 20]),
            len([r for r in results if 20 <= r.percentage < 40]),
            len([r for r in results if 40 <= r.percentage < 60]),
            len([r for r in results if 60 <= r.percentage < 80]),
            len([r for r in results if 80 <= r.percentage <= 100])
        ]
    })